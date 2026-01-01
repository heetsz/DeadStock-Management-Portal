from sqlalchemy.orm import Session
from sqlalchemy import select, func, and_, or_
from typing import Optional, List, Tuple
from datetime import datetime, date
from io import BytesIO
from decimal import Decimal

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Asset, AssetAssignment, Scrap, Lab, Vendor, Category, Teacher
from app.schemas.asset import AssetFilters
from app.services.asset_service import AssetService


class ReportService:
    
    def generate_asset_report(
        self,
        db: Session,
        filters: AssetFilters,
        format: str = 'pdf'
    ) -> Tuple[bytes, str, str]:
        """Generate asset report in requested format"""
        
        # Fetch data
        asset_service = AssetService()
        asset_list = asset_service.get_filtered_assets(
            db, filters, page=1, size=10000  # Get all for report
        )
        
        # Enhance with additional data
        report_data = []
        for asset_response in asset_list.items:
            asset = asset_service.get_asset(db, asset_response.asset_id)
            if not asset:
                continue
            
            # Get category, vendor, lab names
            category_name = None
            if asset.category_id:
                category = db.query(Category).filter(Category.category_id == asset.category_id).first()
                category_name = category.name if category else None
            
            vendor_name = None
            if asset.vendor_id:
                vendor = db.query(Vendor).filter(Vendor.vendor_id == asset.vendor_id).first()
                vendor_name = vendor.vendor_name if vendor else None
            
            lab_info = None
            if asset.lab_id:
                lab = db.query(Lab).filter(Lab.lab_id == asset.lab_id).first()
                if lab:
                    lab_info = f"{lab.lab_name} ({lab.room_number})" if lab.room_number else lab.lab_name
            
            # Get assigned teachers
            assignments = db.query(AssetAssignment).filter(
                and_(
                    AssetAssignment.asset_id == asset.asset_id,
                    AssetAssignment.return_date.is_(None)
                )
            ).all()
            teachers = [a.teacher.name for a in assignments if a.teacher for a in [a]]
            assigned_to = ", ".join(teachers) if teachers else "Not Assigned"
            
            report_data.append({
                'description': asset.description,
                'category': category_name or 'N/A',
                'total_quantity': asset.total_quantity,
                'assigned_quantity': asset_response.active_assigned_quantity,
                'scrapped_quantity': asset_response.total_scrapped_quantity,
                'available_quantity': asset_response.available_quantity,
                'purchase_date': asset.purchase_date.strftime('%d-%b-%Y'),
                'financial_year': asset.financial_year,
                'vendor': vendor_name or 'N/A',
                'original_cost': float(asset.original_total_cost),
                'current_cost': float(asset.current_total_cost),
                'scrap_cost': float(asset.original_total_cost - asset.current_total_cost),
                'lab': lab_info or 'N/A',
                'physical_location': asset.physical_location or 'N/A',
                'assigned_to': assigned_to,
                'remarks': asset.remarks or ''
            })
        
        # Calculate summary
        summary = {
            'total_assets': len(report_data),
            'total_quantity': sum(r['total_quantity'] for r in report_data),
            'total_assigned': sum(r['assigned_quantity'] for r in report_data),
            'total_scrapped': sum(r['scrapped_quantity'] for r in report_data),
            'total_available': sum(r['available_quantity'] for r in report_data),
            'total_original_cost': sum(r['original_cost'] for r in report_data),
            'total_current_cost': sum(r['current_cost'] for r in report_data),
            'total_scrap_cost': sum(r['scrap_cost'] for r in report_data)
        }
        
        # Generate in requested format
        if format == 'pdf':
            file_bytes = self._generate_asset_pdf(report_data, summary, filters)
            filename = f"asset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            content_type = "application/pdf"
        elif format == 'csv':
            file_bytes = self._generate_asset_csv(report_data)
            filename = f"asset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            content_type = "text/csv"
        elif format == 'xlsx':
            file_bytes = self._generate_asset_excel(report_data, summary)
            filename = f"asset_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        return file_bytes, filename, content_type
    
    def _generate_asset_pdf(
        self,
        data: List[dict],
        summary: dict,
        filters: AssetFilters
    ) -> bytes:
        """Generate professional PDF report"""
        buffer = BytesIO()
        
        # Use portrait A4 for better print compatibility
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,  # Changed from landscape to portrait A4
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Header
        college_style = styles['Title']
        college_style.fontSize = 16
        college_style.alignment = 1
        
        elements.append(Paragraph("Sardar Patel Institute of Technology", college_style))
        elements.append(Paragraph("Computer Engineering Department", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        title_style = styles['Heading1']
        title_style.fontSize = 14
        title_style.textColor = colors.HexColor('#1a56db')
        
        elements.append(Paragraph("Asset Master Report", title_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Applied filters
        filter_text = self._format_applied_filters(filters)
        if filter_text:
            filter_style = styles['Normal']
            filter_style.fontSize = 9
            filter_style.textColor = colors.grey
            elements.append(Paragraph(f"<b>Applied Filters:</b> {filter_text}", filter_style))
            elements.append(Spacer(1, 0.15*inch))
        
        # Table - Adjusted for A4 portrait
        table_data = [[
            'Description', 'Category', 'Total', 'Assigned', 'Scrapped', 'Available',
            'Purchase Date', 'FY', 'Vendor', 'Original Cost', 'Current Cost', 'Lab'
        ]]
        
        for row in data:
            table_data.append([
                row['description'][:25],  # Shorter for portrait
                row['category'][:10],
                str(row['total_quantity']),
                str(row['assigned_quantity']),
                str(row['scrapped_quantity']),
                str(row['available_quantity']),
                row['purchase_date'],
                row['financial_year'],
                row['vendor'][:12],
                f"₹{row['original_cost']:,.0f}",
                f"₹{row['current_cost']:,.0f}",
                row['lab'][:10]
            ])
        
        # Adjusted column widths for A4 portrait (total ~7.5 inches)
        table = Table(table_data, colWidths=[
            1.2*inch,  # Description
            0.6*inch,  # Category
            0.4*inch,  # Total
            0.4*inch,  # Assigned
            0.4*inch,  # Scrapped
            0.4*inch,  # Available
            0.6*inch,  # Purchase Date
            0.5*inch,  # FY
            0.7*inch,  # Vendor
            0.7*inch,  # Original Cost
            0.7*inch,  # Current Cost
            0.6*inch   # Lab
        ])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary
        summary_data = [
            ['Summary Statistics', ''],
            ['Total Assets', str(summary['total_assets'])],
            ['Total Quantity Purchased', str(summary['total_quantity'])],
            ['Total Quantity Assigned', str(summary['total_assigned'])],
            ['Total Quantity Scrapped', str(summary['total_scrapped'])],
            ['Total Quantity Available', str(summary['total_available'])],
            ['', ''],
            ['Total Original Cost', f"₹{summary['total_original_cost']:,.2f}"],
            ['Total Scrap Cost', f"₹{summary['total_scrap_cost']:,.2f}"],
            ['Total Current Value', f"₹{summary['total_current_cost']:,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('BACKGROUND', (0, 7), (-1, 9), colors.HexColor('#fef3c7')),
            ('FONTNAME', (0, 7), (-1, 9), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Footer
        footer_style = styles['Normal']
        footer_style.fontSize = 8
        footer_style.textColor = colors.grey
        elements.append(Paragraph(
            f"Generated on: {datetime.now().strftime('%d %B %Y at %I:%M %p')}",
            footer_style
        ))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer.read()
    
    def _generate_asset_csv(self, data: List[dict]) -> bytes:
        """Generate CSV report"""
        buffer = BytesIO()
        writer = csv.writer(buffer)
        
        writer.writerow([
            'Description', 'Category', 'Total Quantity', 'Assigned Quantity',
            'Scrapped Quantity', 'Available Quantity', 'Purchase Date', 'Financial Year',
            'Vendor', 'Original Cost', 'Current Cost', 'Scrap Cost', 'Lab',
            'Physical Location', 'Assigned To', 'Remarks'
        ])
        
        for row in data:
            writer.writerow([
                row['description'], row['category'], row['total_quantity'],
                row['assigned_quantity'], row['scrapped_quantity'], row['available_quantity'],
                row['purchase_date'], row['financial_year'], row['vendor'],
                row['original_cost'], row['current_cost'], row['scrap_cost'],
                row['lab'], row['physical_location'], row['assigned_to'], row['remarks']
            ])
        
        buffer.seek(0)
        return buffer.read()
    
    def _generate_asset_excel(self, data: List[dict], summary: dict) -> bytes:
        """Generate Excel report"""
        buffer = BytesIO()
        wb = openpyxl.Workbook()
        
        ws_data = wb.active
        ws_data.title = "Asset Data"
        
        headers = [
            'Description', 'Category', 'Total Qty', 'Assigned', 'Scrapped', 'Available',
            'Purchase Date', 'Financial Year', 'Vendor', 'Original Cost', 'Current Cost',
            'Scrap Cost', 'Lab', 'Physical Location', 'Assigned To', 'Remarks'
        ]
        
        header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_num, row_data in enumerate(data, 2):
            ws_data.cell(row=row_num, column=1, value=row_data['description'])
            ws_data.cell(row=row_num, column=2, value=row_data['category'])
            ws_data.cell(row=row_num, column=3, value=row_data['total_quantity'])
            ws_data.cell(row=row_num, column=4, value=row_data['assigned_quantity'])
            ws_data.cell(row=row_num, column=5, value=row_data['scrapped_quantity'])
            ws_data.cell(row=row_num, column=6, value=row_data['available_quantity'])
            ws_data.cell(row=row_num, column=7, value=row_data['purchase_date'])
            ws_data.cell(row=row_num, column=8, value=row_data['financial_year'])
            ws_data.cell(row=row_num, column=9, value=row_data['vendor'])
            ws_data.cell(row=row_num, column=10, value=row_data['original_cost'])
            ws_data.cell(row=row_num, column=11, value=row_data['current_cost'])
            ws_data.cell(row=row_num, column=12, value=row_data['scrap_cost'])
            ws_data.cell(row=row_num, column=13, value=row_data['lab'])
            ws_data.cell(row=row_num, column=14, value=row_data['physical_location'])
            ws_data.cell(row=row_num, column=15, value=row_data['assigned_to'])
            ws_data.cell(row=row_num, column=16, value=row_data['remarks'])
            
            if row_num % 2 == 0:
                fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
                for col in range(1, 17):
                    ws_data.cell(row=row_num, column=col).fill = fill
        
        # Auto-adjust column widths
        for col in ws_data.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column].width = adjusted_width
        
        # Summary sheet
        ws_summary = wb.create_sheet(title="Summary")
        summary_data = [
            ['Asset Report Summary', ''],
            ['', ''],
            ['Metric', 'Value'],
            ['Total Assets', summary['total_assets']],
            ['Total Quantity Purchased', summary['total_quantity']],
            ['Total Quantity Assigned', summary['total_assigned']],
            ['Total Quantity Scrapped', summary['total_scrapped']],
            ['Total Quantity Available', summary['total_available']],
            ['', ''],
            ['Financial Summary', ''],
            ['Total Original Cost', summary['total_original_cost']],
            ['Total Scrap Cost', summary['total_scrap_cost']],
            ['Total Current Value', summary['total_current_cost']]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color='FFFFFF', size=14)
                elif row_num == 3:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                elif row_num == 10:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    
    def _format_applied_filters(self, filters: AssetFilters) -> str:
        """Format applied filters into readable text"""
        filter_parts = []
        
        if filters.financial_year:
            filter_parts.append(f"FY: {filters.financial_year}")
        
        if filters.lab_id:
            filter_parts.append(f"Lab ID: {filters.lab_id}")
        
        if filters.category_id:
            filter_parts.append(f"Category ID: {filters.category_id}")
        
        if filters.issued_status:
            status_map = {
                'issued_only': 'Issued Only',
                'not_issued': 'Not Issued',
                'partially_issued': 'Partially Issued'
            }
            filter_parts.append(f"Status: {status_map.get(filters.issued_status)}")
        
        if filters.scrap_status:
            scrap_map = {
                'scrapped_only': 'Scrapped Only',
                'exclude_scrapped': 'Exclude Scrapped'
            }
            filter_parts.append(f"Scrap: {scrap_map.get(filters.scrap_status)}")
        
        if filters.search:
            filter_parts.append(f"Search: '{filters.search}'")
        
        return " | ".join(filter_parts) if filter_parts else "No filters applied"
    
    def generate_assignment_report(
        self,
        db: Session,
        asset_id: Optional[str] = None,
        teacher_id: Optional[str] = None,
        lab_id: Optional[str] = None,
        category_id: Optional[str] = None,
        active_only: Optional[bool] = None,
        assignment_date_from: Optional[date] = None,
        assignment_date_to: Optional[date] = None,
        format: str = 'pdf'
    ) -> Tuple[bytes, str, str]:
        """Generate assignment report in requested format"""
        
        from app.services.assignment_service import AssignmentService
        assignment_service = AssignmentService()
        
        # Get assignments with filters
        assignments = assignment_service.get_assignments(
            db, asset_id=asset_id, teacher_id=teacher_id, active_only=active_only if active_only is not None else False
        )
        
        # Apply additional filters
        report_data = []
        for assignment in assignments:
            asset = db.query(Asset).filter(Asset.asset_id == assignment.asset_id).first()
            if not asset:
                continue
            
            # Filter by lab_id if provided
            if lab_id and asset.lab_id != lab_id:
                continue
            
            # Filter by category_id if provided
            if category_id and asset.category_id != category_id:
                continue
            
            # Filter by date range
            if assignment_date_from and assignment.assignment_date < assignment_date_from:
                continue
            if assignment_date_to and assignment.assignment_date > assignment_date_to:
                continue
            
            # Get related data
            teacher_name = assignment.teacher.name if assignment.teacher else 'N/A'
            category_name = asset.category.name if asset.category else 'N/A'
            lab_name = asset.lab.lab_name if asset.lab else 'N/A'
            vendor_name = asset.vendor.vendor_name if asset.vendor else 'N/A'
            
            # Calculate assigned cost
            per_unit_cost = float(asset.original_total_cost) / asset.total_quantity if asset.total_quantity > 0 else 0
            assigned_cost = per_unit_cost * assignment.assigned_quantity
            
            report_data.append({
                'assignment_id': assignment.assignment_id,
                'asset_description': asset.description,
                'category': category_name,
                'teacher_name': teacher_name,
                'assigned_quantity': assignment.assigned_quantity,
                'assignment_date': assignment.assignment_date.strftime('%d-%b-%Y'),
                'return_date': assignment.return_date.strftime('%d-%b-%Y') if assignment.return_date else 'Active',
                'status': 'Active' if assignment.return_date is None else 'Returned',
                'lab': lab_name,
                'vendor': vendor_name,
                'financial_year': asset.financial_year,
                'purchase_date': asset.purchase_date.strftime('%d-%b-%Y'),
                'original_cost': float(asset.original_total_cost),
                'assigned_cost': assigned_cost,
                'current_location': assignment.current_location or 'N/A',
                'remarks': assignment.remarks or ''
            })
        
        # Calculate summary
        summary = {
            'total_assignments': len(report_data),
            'active_assignments': len([r for r in report_data if r['status'] == 'Active']),
            'returned_assignments': len([r for r in report_data if r['status'] == 'Returned']),
            'total_assigned_quantity': sum(r['assigned_quantity'] for r in report_data),
            'total_assigned_cost': sum(r['assigned_cost'] for r in report_data),
            'active_assigned_cost': sum(r['assigned_cost'] for r in report_data if r['status'] == 'Active')
        }
        
        # Generate in requested format
        if format == 'pdf':
            file_bytes = self._generate_assignment_pdf(report_data, summary, {
                'asset_id': asset_id, 'teacher_id': teacher_id, 'lab_id': lab_id,
                'category_id': category_id, 'active_only': active_only
            })
            filename = f"assignment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            content_type = "application/pdf"
        elif format == 'csv':
            file_bytes = self._generate_assignment_csv(report_data)
            filename = f"assignment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            content_type = "text/csv"
        elif format == 'xlsx':
            file_bytes = self._generate_assignment_excel(report_data, summary)
            filename = f"assignment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        return file_bytes, filename, content_type
    
    def _generate_assignment_pdf(
        self,
        data: List[dict],
        summary: dict,
        filters: dict
    ) -> bytes:
        """Generate professional PDF report for assignments"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Header
        college_style = styles['Title']
        college_style.fontSize = 16
        college_style.alignment = 1
        
        elements.append(Paragraph("Sardar Patel Institute of Technology", college_style))
        elements.append(Paragraph("Computer Engineering Department", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        title_style = styles['Heading1']
        title_style.fontSize = 14
        title_style.textColor = colors.HexColor('#1a56db')
        
        elements.append(Paragraph("Assignment Report", title_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Applied filters
        filter_parts = []
        if filters.get('lab_id'):
            filter_parts.append(f"Lab ID: {filters['lab_id']}")
        if filters.get('teacher_id'):
            filter_parts.append(f"Teacher ID: {filters['teacher_id']}")
        if filters.get('category_id'):
            filter_parts.append(f"Category ID: {filters['category_id']}")
        if filters.get('active_only'):
            filter_parts.append("Status: Active Only")
        
        if filter_parts:
            filter_style = styles['Normal']
            filter_style.fontSize = 9
            filter_style.textColor = colors.grey
            elements.append(Paragraph(f"<b>Applied Filters:</b> {' | '.join(filter_parts)}", filter_style))
            elements.append(Spacer(1, 0.15*inch))
        
        # Table
        table_data = [[
            'Asset Description', 'Category', 'Teacher', 'Qty', 'Assigned Date',
            'Return Date', 'Status', 'Lab', 'Assigned Cost'
        ]]
        
        for row in data:
            table_data.append([
                row['asset_description'][:20],
                row['category'][:10],
                row['teacher_name'][:15],
                str(row['assigned_quantity']),
                row['assignment_date'],
                row['return_date'][:10],
                row['status'],
                row['lab'][:10],
                f"₹{row['assigned_cost']:,.0f}"
            ])
        
        table = Table(table_data, colWidths=[
            1.2*inch, 0.7*inch, 0.9*inch, 0.4*inch, 0.7*inch,
            0.7*inch, 0.5*inch, 0.6*inch, 0.8*inch
        ])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary
        summary_data = [
            ['Summary Statistics', ''],
            ['Total Assignments', str(summary['total_assignments'])],
            ['Active Assignments', str(summary['active_assignments'])],
            ['Returned Assignments', str(summary['returned_assignments'])],
            ['Total Assigned Quantity', str(summary['total_assigned_quantity'])],
            ['', ''],
            ['Total Assigned Cost', f"₹{summary['total_assigned_cost']:,.2f}"],
            ['Active Assigned Cost', f"₹{summary['active_assigned_cost']:,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('BACKGROUND', (0, 5), (-1, 7), colors.HexColor('#fef3c7')),
            ('FONTNAME', (0, 5), (-1, 7), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Footer
        footer_style = styles['Normal']
        footer_style.fontSize = 8
        footer_style.textColor = colors.grey
        elements.append(Paragraph(
            f"Generated on: {datetime.now().strftime('%d %B %Y at %I:%M %p')}",
            footer_style
        ))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer.read()
    
    def _generate_assignment_csv(self, data: List[dict]) -> bytes:
        """Generate CSV report for assignments"""
        buffer = BytesIO()
        writer = csv.writer(buffer)
        
        writer.writerow([
            'Assignment ID', 'Asset Description', 'Category', 'Teacher Name', 'Assigned Quantity',
            'Assignment Date', 'Return Date', 'Status', 'Lab', 'Vendor', 'Financial Year',
            'Purchase Date', 'Original Cost', 'Assigned Cost', 'Current Location', 'Remarks'
        ])
        
        for row in data:
            writer.writerow([
                row['assignment_id'], row['asset_description'], row['category'],
                row['teacher_name'], row['assigned_quantity'], row['assignment_date'],
                row['return_date'], row['status'], row['lab'], row['vendor'],
                row['financial_year'], row['purchase_date'], row['original_cost'],
                row['assigned_cost'], row['current_location'], row['remarks']
            ])
        
        buffer.seek(0)
        return buffer.read()
    
    def _generate_assignment_excel(self, data: List[dict], summary: dict) -> bytes:
        """Generate Excel report for assignments"""
        buffer = BytesIO()
        wb = openpyxl.Workbook()
        
        ws_data = wb.active
        ws_data.title = "Assignment Data"
        
        headers = [
            'Assignment ID', 'Asset Description', 'Category', 'Teacher Name', 'Assigned Qty',
            'Assignment Date', 'Return Date', 'Status', 'Lab', 'Vendor', 'Financial Year',
            'Purchase Date', 'Original Cost', 'Assigned Cost', 'Current Location', 'Remarks'
        ]
        
        header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_num, row_data in enumerate(data, 2):
            ws_data.cell(row=row_num, column=1, value=row_data['assignment_id'])
            ws_data.cell(row=row_num, column=2, value=row_data['asset_description'])
            ws_data.cell(row=row_num, column=3, value=row_data['category'])
            ws_data.cell(row=row_num, column=4, value=row_data['teacher_name'])
            ws_data.cell(row=row_num, column=5, value=row_data['assigned_quantity'])
            ws_data.cell(row=row_num, column=6, value=row_data['assignment_date'])
            ws_data.cell(row=row_num, column=7, value=row_data['return_date'])
            ws_data.cell(row=row_num, column=8, value=row_data['status'])
            ws_data.cell(row=row_num, column=9, value=row_data['lab'])
            ws_data.cell(row=row_num, column=10, value=row_data['vendor'])
            ws_data.cell(row=row_num, column=11, value=row_data['financial_year'])
            ws_data.cell(row=row_num, column=12, value=row_data['purchase_date'])
            ws_data.cell(row=row_num, column=13, value=row_data['original_cost'])
            ws_data.cell(row=row_num, column=14, value=row_data['assigned_cost'])
            ws_data.cell(row=row_num, column=15, value=row_data['current_location'])
            ws_data.cell(row=row_num, column=16, value=row_data['remarks'])
            
            if row_num % 2 == 0:
                fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
                for col in range(1, 17):
                    ws_data.cell(row=row_num, column=col).fill = fill
        
        # Auto-adjust column widths
        for col in ws_data.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column].width = adjusted_width
        
        # Summary sheet
        ws_summary = wb.create_sheet(title="Summary")
        summary_data = [
            ['Assignment Report Summary', ''],
            ['', ''],
            ['Metric', 'Value'],
            ['Total Assignments', summary['total_assignments']],
            ['Active Assignments', summary['active_assignments']],
            ['Returned Assignments', summary['returned_assignments']],
            ['Total Assigned Quantity', summary['total_assigned_quantity']],
            ['', ''],
            ['Financial Summary', ''],
            ['Total Assigned Cost', summary['total_assigned_cost']],
            ['Active Assigned Cost', summary['active_assigned_cost']]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:
                    cell.font = Font(bold=True, size=14)
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color='FFFFFF', size=14)
                elif row_num == 3:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                elif row_num == 9:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

